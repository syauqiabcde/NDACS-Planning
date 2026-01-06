#%%
import numpy as np
from pyomo.util.infeasible import log_infeasible_constraints
from pyomo.environ import *
import pandas as pd
import openpyxl as xl 
import time
import matplotlib.pyplot as plt
from scipy.interpolate import RegularGridInterpolator
from amplpy import modules

print("Start")
start = time.time()

# Case name
case_name = 'Robust Base case'
print(f"case name: {case_name}")
# General data and asumption
ppm_to_mass = 2.13 * 1e9      # Conversion from ppm to mass CO2, 1 ppm of CO2 = 2.13 giga ton of CO2
percent_allowable_investment = 0.05
plant_lifetime = 30
uranium_consumption = 1 # in kg uranium/capacity/year 
uranium_reserve = 8e9 # uranium reserve in the world
water_consumption = 1 # in m3/capacity/year
land_consumption = 1 # in m2/capacity
discount_rate = 0.1 # 10% of discount rate
fixed_om = 0.087 # 8.7% fixed O&M cost of Nuclear DACCS plant
plant_cf = 0.7

scenario = 'IPCC middle'
ppm_limit_scenario = '1.5 C'
ppm_limit ={'1.5 C': (465.0, 411.0),
           '2 C': (505.0, 480.0)}

# Getting data from excel
wb = xl.load_workbook('data_5years.xlsx') 
sheet_name = wb.sheetnames
df = {sheet: pd.read_excel('data_5years.xlsx', sheet_name=sheet) for sheet in sheet_name}
years = df['capex'].iloc[:,0].values
countries = df['env_condition'].iloc[:,0].values
nuclear_allowed = df['nuclear_allowed'].iloc[:,:].values.reshape(-1)
ccs_included = df['ccs_included'].iloc[:,:].values.reshape(-1)
non_NPT = df['non-NPT'].iloc[:,:].values.reshape(-1)
interval = years[1] - years[0]

# Weather data
wb_weather = xl.load_workbook('weather_data.xlsx') 
sheet_name_weather = wb_weather.sheetnames
weather_years = [2011, 2013, 2014, 2015, 2017, 2018, 2019, 2021, 2022, 2023]
num_hours = pd.read_excel('weather_data.xlsx', sheet_name=f"temperature_data_{weather_years[0]}").shape[1] - 1  

num_periods = len(years)
num_countries = len(countries)

# Define model
model = ConcreteModel()

# Sets
model.i = Set(initialize=countries)
model.i_nuclear_allowed = Set(initialize=nuclear_allowed)
model.i_ccs_included = Set(initialize=ccs_included)
model.i_non_NPT = Set(initialize=non_NPT)
model.t = Set(initialize=years)

# Parameters
duty = df['energy_duty'].iloc[:,1:].values
temperature = np.linspace(0, 40, 17)
rh = np.linspace(0, 100, 21)
co2_tax = df['co2_tax'].iloc[:,1:].values
co2_storage_capacity = df['co2_storage_capacity'].iloc[:,1].values * 1e9
gdp = df['gdp'].iloc[:,1:].values
plant_capex = df['capex'].iloc[:,1].values
ccs_readiness_index = df['ccs_readiness'].iloc[:,1].values
nuclear_readiness_index = df['nuclear_readiness'].iloc[:,1].values
nuclear_support = df['nuclear_perception'].iloc[:,1].values
nuclear_oppose = df['nuclear_perception'].iloc[:,2].values
water_reserve = df['resource_reserve'].iloc[:,1].values
land_reserve = df['resource_reserve'].iloc[:,2].values

interpolator = RegularGridInterpolator((rh,temperature), duty, bounds_error=False, fill_value=None)

energy_duty_data = []

for i, country in enumerate(countries):
    weather_data = []
    for year in weather_years:
        # Load temperature and humidity for the year
        temp_df = pd.read_excel('weather_data.xlsx', sheet_name=f"temperature_data_{year}")
        rh_df   = pd.read_excel('weather_data.xlsx', sheet_name=f"humidity_data_{year}")
        
        temp_row = temp_df.iloc[i, 1:].values 
        rh_row   = rh_df.iloc[i, 1:].values
        weather_row = np.stack((temp_row, rh_row))
        weather_data.append(weather_row)

    weather_array = np.stack(weather_data, axis=2)

    T  = weather_array[0] 
    RH = weather_array[1]  

    points = np.column_stack([RH.ravel(), T.ravel()])  
    duty_all = interpolator(points)   
    duty_all = duty_all.reshape(num_hours, -1)                
    duty_max_per_hour = duty_all.max(axis=1)
    energy_duty_data.append(duty_max_per_hour)

duty_data_array = np.average(np.stack(energy_duty_data), axis=1)  

elc_cost = 0.0154 # $/MJ
heat_cost = 0.0067 # $/MJ
heat_share = 0.87   # percentage of heat energy requirement compared to total energy requirement / elc_share = 1-heat_share

duty_data = {(countries[i]): duty_data_array[i] for i in range(num_countries)}
co2_tax_data = {(countries[i], years[t]): co2_tax[i, t] for i in range(num_countries) for t in range(num_periods)}
CO2_storage_data = {countries[i]: co2_storage_capacity[i] for i in range(num_countries)}
gdp_data = {(countries[i], years[t]): gdp[i, t] for i in range(num_countries) for t in range(num_periods)}
capex_data = {years[t]: plant_capex[t] for t in range(num_periods)}
ccs_readiness_data = {countries[i]: ccs_readiness_index[i] for i in range(num_countries)}
nuclear_readiness_data = {countries[i]: nuclear_readiness_index[i] for i in range(num_countries)}
nuclear_support_data = {countries[i]: nuclear_support[i] for i in range(num_countries)}
nuclear_oppose_data = {countries[i]: nuclear_oppose[i] for i in range(num_countries)}
water_reserve_data = {countries[i]: water_reserve[i] for i in range(num_countries)}
land_reserve_data = {countries[i]: land_reserve[i] for i in range(num_countries)}

model.energy_duty = Param(model.i, initialize=duty_data, within=Reals)
model.co2_tax = Param(model.i, model.t, initialize=co2_tax_data, within=Reals)
model.CO2_storage = Param(model.i, initialize=CO2_storage_data , within=Reals)
model.gdp = Param(model.i, model.t, initialize=gdp_data , within=Reals)
model.capex = Param(model.t, initialize=capex_data , within=Reals)
model.ccs_readiness = Param(model.i, initialize=ccs_readiness_data , within=Reals)
model.nuclear_readiness = Param(model.i, initialize=nuclear_readiness_data , within=Integers)
model.nuclear_support = Param(model.i, initialize=nuclear_support_data , within=Reals)
model.nuclear_oppose = Param(model.i, initialize=nuclear_oppose_data , within=Reals)    
model.water_reserve = Param(model.i, initialize=water_reserve_data , within=Reals)
model.land_reserve = Param(model.i, initialize=land_reserve_data , within=Reals)

print("Parameters initialization finished.")

#%% Model construction

# Decision Variables
model.new = Var(model.i, model.t, within=NonNegativeReals, bounds=(0.0, 1.0 * 1e9))
model.co2_ppm = Var(model.t, within=NonNegativeReals)       # not really a decision variables just a way to determine the co2 ppm based on the previous value of co2 ppm

# Expressions
def retirement(model, i, t):
        retirement_period = t - plant_lifetime  # Find the period that corresponds to retirement period
        if retirement_period in model.t:
            return model.new[i, retirement_period]  # Retire what has been exceed lifetime
        else:
            return 0
        
def capacity_expansion(model, i, t):
    if t == min(model.t):
        return model.new[i,t]
    else:
        return model.capacity[i,t-interval] + model.new[i,t] - model.retired[i,t]
 
def IPCC_prediction(t, scenario='IPCC middle'):        # return the CO2 concentration in ppm based on IPCC prediction 
    if scenario == 'IPCC middle':
        return -0.00804 * t**2 + 34.83214 * t - 37159.42857
    elif scenario == 'IPCC conservative':
        return 0.05091 * t**2 - 203.18805 * t + 203146.33017 

def captured_co2(model, i, t):   # hourly CO2 captured
    return model.capacity[i,t] * plant_cf * interval

model.retired = Expression(model.i, model.t, rule=retirement)
model.capacity = Expression(model.i, model.t, rule=capacity_expansion)
model.captured_co2 = Expression(model.i, model.t, rule=captured_co2)

# Objective Function
def tot_cost(model):
    # return sum(((model.cost_table[i, t] - model.co2_tax[i, t]) / (1+discount_rate)**(t-min(model.t))) * model.capacity[i, t] for i in model.i for t in model.t)
    return sum(((model.capex[t] * model.new[i,t]                        # Investment cost
                + model.capex[t] * model.capacity[i,t] * fixed_om       # Fixed OM cost of Nuclear DACCS plant
                + model.energy_duty[i] * 1000 * model.captured_co2[i,t] * (heat_share * heat_cost + (1-heat_share) * elc_cost))  # energy cost, multiply 1000 due to conversion from MJ/kg to MJ/ton
                - model.co2_tax[i, t] * model.captured_co2[i,t])        # CO2 tax benefit
                / (1+discount_rate)**(t-min(model.t))                     # Discount rate  
                for i in model.i for t in model.t )       
    
model.obj = Objective(rule=tot_cost, sense=minimize)

# Mandatory Constraints
def CO2_ppm_limit(model, t):
    if t == max(model.t):
        return model.co2_ppm[t] <= ppm_limit[ppm_limit_scenario][1]        # the co2 ppm in 2100 must lower than 411 ppm (Paris agreement 1.5 C scenario) 
    else:
        return model.co2_ppm[t] <= ppm_limit[ppm_limit_scenario ][0]        # the co2 ppm in all years must lower than 465 ppm (Paris agreement 1.5 C scenario) 

def CO2_storage(model, i):
    return sum(model.captured_co2[i,t] for t in model.t) <= model.CO2_storage[i]

def co2_ppm_rule(model, t):
    if t == min(model.t):  # Initial CO2 concentration at the first period
        captured = sum(model.captured_co2[i, t] for i in model.i)
        return model.co2_ppm[t] == IPCC_prediction(t, scenario) - captured / ppm_to_mass
    else:
        generated_co2 = (IPCC_prediction(t, scenario) - IPCC_prediction(t-interval, scenario)) * ppm_to_mass
        captured = sum(model.captured_co2[i, t] for i in model.i)
        return model.co2_ppm[t] == model.co2_ppm[t-interval] + (generated_co2 - captured) / ppm_to_mass

model.ppm_limit = Constraint(model.t, rule=CO2_ppm_limit)
model.storage_constraint = Constraint(model.i, rule=CO2_storage)
model.co2_atm_balance = Constraint(model.t, rule=co2_ppm_rule)

# Optional Constraints (please comment in which constraint you dont want to apply)
def financing_constrant(model, i, t):
    return model.new[i, t] * model.capex[t] <= percent_allowable_investment * model.gdp[i, t] * 1000

def profit_constraint(model, i, t):
    return model.captured_co2[i,t] * model.co2_tax[i,t] >= model.capacity[i,t] * model.cost_table[i,t]

def ccs_readiness_constraint(model, i):
    if model.ccs_readiness[i] < 50:
        return model.new[i,t] == 0
    return Constraint.Skip

def nuclear_readiness_constraint(model, i):
    if model.nuclear_readiness[i] < 3:
        return model.new[i,t] == 0
    return Constraint.Skip

def countries_allowed_constraint(model, i, t):
    if i not in model.i_nuclear_allowed.data():
        return model.new[i,t] == 0
    return Constraint.Skip

def countries_NPT(model, i, t):
    if i in model.i_non_NPT.data():
        return model.new[i,t] == 0
    return Constraint.Skip

def public_perception_constraint(model, i, t):
    if model.nuclear_support[i] < model.nuclear_oppose[i]:
        return model.new[i,t] == 0
    return Constraint.Skip

def uranium_reserve_constraint(model):
    return sum(model.capacity[i,t] for i in model.i for t in model.t) * uranium_consumption <= uranium_reserve

def water_reserve_constraint(model, i):
    return sum(model.capacity[i,t] for t in model.t) * water_consumption <= model.water_reserve[i]

def land_reserve_constraint(model, i):
    return sum(model.capacity[i,t] for t in model.t) * land_consumption <= model.land_reserve[i]

def ccs_included_constraint(model, i, t):
    if i not in model.i_ccs_included:
        return model.new[i,t] == 0
    return Constraint.Skip

# Economic aspect
# model.financing_constraint = Constraint(model.i, model.t, rule=financing_constrant)

# Technological readiness
# model.ccs_readiness_constraint = Constraint(model.i, rule=ccs_readiness_constraint)
# model.nuclear_readiness_constraint = Constraint(model.i, rule=nuclear_readiness_constraint)

# Socio-political aspect
# model.nuclear_allowed = Constraint(model.i, model.t, rule=countries_allowed_constraint)
# model.NPT_constratint = Constraint(model.i, model.t, rule=countries_NPT)
# model.nuclear_perception_constraint = Constraint(model.i, model.t, rule=public_perception_constraint)

# Technical aspect
# model.uranium_reserve_constraint = Constraint(rule=uranium_reserve_constraint)
# model.water_reserve_constraint = Constraint(model.i, rule=water_reserve_constraint)
# model.land_reserve_constraint = Constraint(model.i, rule=land_reserve_constraint)

# Political wilingness aspect
# model.ccs_included_constraint = Constraint(model.i, model.t, rule=ccs_included_constraint)

print("Model construction finished.")
#%% Solver

print("Solving the model...")
# Solver
solver = SolverFactory('glpk', executable='C:\\w64\\glpsol.exe')
# solver = SolverFactory('ipopt', executable='C:\\w64\\ipopt.exe')
# solver.options['max_iter'] = 5000
result = solver.solve(model)

print(f'The solver status is {result.solver.status}')
print(f'The termination condition is {result.solver.termination_condition}')

if result.solver.termination_condition == 'infeasible':
    print("The model is infeasible. Please check the constraints and data.")
    log_infeasible_constraints(model)

#%% Result export

if result.solver.termination_condition == 'optimal':
    print(f'the total present cost is: {model.obj()/1e9:.2f} Billion USD')

    # Getting result
    result_dict = {
    "New Capacity": np.zeros((num_countries, num_periods)),
    "Capacity": np.zeros((num_countries, num_periods)),
    "Investment Required": np.zeros((num_countries, num_periods)),
    "CO2 Storage Level": np.zeros(num_countries),
    "CO2 ppm": np.zeros(num_periods)
    }
    
    unit_dict = {
        "New Capacity": 'million CO2 per year',
        "Capacity": 'million CO2 per year',
        "Investment Required": 'million USD',
        "CO2 Storage Level": '%',
        "CO2 ppm": 'ppm'
    }

    for x, i in enumerate(model.i):
        for y, t in enumerate(model.t):
            result_dict['New Capacity'][x,y] = model.new[i, t].value / 1e6         # New capacity in million CO2/year
            result_dict['Capacity'][x, y] = model.capacity[i, t]() / 1e6      # Capacity in million CO2/year
            result_dict['Investment Required'][x,y] = (model.new[i, t].value * capex_data[t]) / 1e6   # Investment required in million USD

            if y == 0:
                result_dict['CO2 Storage Level'][x] = sum(model.captured_co2[i,t]() for t in model.t) / max(1e-10, model.CO2_storage[i])       # co2 storage level (100% means full 0% empty)

            if x == 0:
                result_dict['CO2 ppm'][y] = model.co2_ppm[t]()               # co2 concentration in atmosphere in ppm

    for key in ["New Capacity", "Capacity", "Investment Required"]:      # # data that belong to i and t set
        result_dict[key] = np.column_stack((countries, result_dict[key]))  
        result_dict[key] = pd.DataFrame(result_dict[key], columns=['Country'] + list(years))

    for key in ["CO2 Storage Level"]:     # data that only belong to i set
        result_dict[key] = np.column_stack((countries, result_dict[key]))  
        result_dict[key] = pd.DataFrame(result_dict[key], columns=['Country', key])

    for key in ["CO2 ppm"]:     # data that only belong to t set
        result_dict[key] = np.column_stack((years, result_dict[key]))  
        result_dict[key] = pd.DataFrame(result_dict[key], columns=['Year', key])

    with pd.ExcelWriter(f"Result/result {case_name}.xlsx") as writer:
        for sheet_name, df in result_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Excel file created successfully! file name: result {case_name}.xlsx")
  
end = time.time()

print(f'total runtime: {end-start:.2f} s')