import pandas as pd
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
import geopandas as gpd
import openpyxl as xl 
import seaborn as sns

plt.rcParams.update({
    "font.size": 8,
    "font.family": "sans-serif",
    "font.sans-serif": ["Helvetica", "Arial", "DejaVu Sans"],
    "axes.labelsize": 10,
    "axes.titlesize": 10,
    "legend.fontsize": 10,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "lines.linewidth": 1.8,
    "axes.linewidth": 0.8,
    "savefig.dpi": 500,
})

class Plotting:
    def __init__(self, optim_scenario, co2_scenario=None):
        self.optim_scenario = optim_scenario
        self.co2_scenario = co2_scenario
        self.sum_params = ['Capacity', 
                           'New Capacity', 
                           'Nuclear Capacity', 
                           'New Nuclear Capacity', 
                           'Investment Required']
        
        self.years = np.arange(2030, 2101, 5)
        self.name_fix = {
            "United States of America": "United States",
            "Russian Federation": "Russia",
            "Viet Nam": "Vietnam",
            "Iran (Islamic Republic of)": "Iran",
            "Korea, Republic of": "South Korea",
            "Korea, Democratic People's Republic of": "North Korea",
            "Côte d'Ivoire": "Ivory Coast",
            "Dem. Rep. Congo": "DR Congo",
            "Republic of the Congo": "Congo",
            "Lao People's Democratic Republic": "Laos",
            "Syrian Arab Republic": "Syria",
            "Bolivia (Plurinational State of)": "Bolivia",
            "Venezuela (Bolivarian Republic of)": "Venezuela",
            "Tanzania, United Republic of": "Tanzania",
            'Central African Rep.': 'Central African Republic',
            'Eq. Guinea': 'Equatorial Guinea',
            'Dominican Rep.': 'Dominican Republic',
            'Czechia': 'Czech Republic',
            'N. Cyprus': 'Cyprus',
            'Trinidad and Tobago': 'Trinidad & Tobago',
            'S. Sudan': 'South Sudan',
            'Somaliland': 'Somalia',
            'Bosnia and Herz.': 'Bosnia & Herzegovina',
            'Solomon Is.': 'Solomon Islands',
            'eSwatini': 'Eswatini'
        }

        self.file_name = {0: rf'Result\Result base case {self.co2_scenario}.xlsx',
                    1: rf'Result\Result economic limitation {self.co2_scenario}.xlsx',
                    2: rf'Result\Result technological limitation {self.co2_scenario}.xlsx',
                    3: rf'Result\Result socio-political limitation {self.co2_scenario}.xlsx',
                    4: rf'Result\Result resource limitation {self.co2_scenario}.xlsx',
                    5: rf'Result\Result all limitations {self.co2_scenario}.xlsx'
                    }

        self.scenario_names = {0: "Base case",
                        1: "Economic limitation",
                        2: "Technological limitation",
                        3: "Socio-political limitation",
                        4: "Resource limitation",
                        5: "All limitations"
                        }

        self.labels = {'Capacity'    : 'DAC Capacity (GtCO$_{2}$/year)',
                'New Capacity': 'New DAC Capacity (GtCO$_{2}$/year)',
                'Nuclear Capacity': 'Nuclear Capacity (GW)',
                'New Nuclear Capacity': 'New Nuclear Capacity (GW)',
                'Investment Required': 'Investment Required (Billion USD$_{2030}$)',
                'Export electricity': 'Export Electricity (TWh/year)',
                'Land consumption': r'Land Consumption (% of total land area)',
                'Water consumption': r'Water Consumption (% of total water consumption)',
                'CO2 ppm': 'Atmospheric CO$_{2}$ Concentration (ppm)',
                'Investment Required': 'Investment Required (Billion USD$_{2030}$)'
                }

        self.n_scenarios = len(self.scenario_names)

        # Load geographical data
        self.world = gpd.read_file(r"natural earth\ne_110m_admin_0_countries.shp")
        self.region_map = pd.read_csv("region_map.csv")
        self.world["NAME"] = self.world["NAME"].replace(self.name_fix)
        self.world = self.world.merge(
            self.region_map,
            left_on="NAME",
            right_on="country",
            how="left"
        )
        self.regions = self.world.dissolve(by="region", as_index=False)

        # Get the data
        wb = xl.load_workbook(self.file_name[self.optim_scenario]) 
        sheet_name = wb.sheetnames
        self.df = {sheet: pd.read_excel(self.file_name[self.optim_scenario], sheet_name=sheet) for sheet in sheet_name}

    def plot_map(self, parameter, year, inpercent=False):
        if inpercent:
            results = self.df[parameter] * 100
        else:
            results = self.df[parameter]

        results_long = results.melt(
            id_vars="Country",
            var_name="year",
            value_name=parameter
        )
        results_long.rename(columns={"Country": "region"}, inplace=True)

        results_year = results_long[
            results_long["year"] == year
        ]

        regions_plot = self.regions.merge(
            results_year,
            on="region",
            how="left"
        )

        fig, ax = plt.subplots(1, 1, figsize=(14, 8), dpi = 600)

        ax.axis("off")
        vmax = results_long[parameter].max()
        vmin = results_long[parameter].min()
        regions_plot.plot(
            column=parameter,
            ax=ax,
            cmap="YlGnBu",
            legend=False,
            edgecolor="black",
            linewidth=0.6,
            vmin= vmin,
            vmax= vmax,
            missing_kwds={
                "color": "lightgrey",
                "label": "No data"
            }
        )
        sm = plt.cm.ScalarMappable(
            cmap="YlGnBu",
            norm=plt.Normalize(vmin=0, vmax=vmax)
        )
        sm._A = []

        cbar = plt.colorbar(sm, ax=ax, shrink=0.6)
        cbar.set_label(self.labels[parameter], fontsize=12, y=0.5)
        ax.text(0.0,1.0, f'Year: {year}', horizontalalignment='left', verticalalignment='top', transform=ax.transAxes, fontsize=12)
        fig.savefig(f'figures\{self.co2_scenario}_{parameter}_{year}_{self.scenario_names[self.optim_scenario]}.png', 
                    dpi=500, 
                    bbox_inches='tight')
        plt.close()

    def plot_line(self, parameter: str, plot_cumsum: bool):
        fig, ax = plt.subplots(figsize=(6, 4), dpi=600)
        if plot_cumsum:
            fig1, ax1 = plt.subplots(figsize=(6, 4), dpi=600)
        colors = sns.color_palette("tab10", n_colors=len(self.scenario_names.keys())+1)

        for optim_scenario in self.scenario_names.keys():
            wb = xl.load_workbook(self.file_name[optim_scenario]) 
            sheet_name = wb.sheetnames
            df = {sheet: pd.read_excel(self.file_name[optim_scenario], sheet_name=sheet) for sheet in sheet_name}

            results = df[parameter]
            if parameter in self.sum_params:
                values = results.iloc[:,1:].sum(axis=0)
                cumsum = np.cumsum(values)
            else:
                values = results.iloc[:,1:].values
            
            if plot_cumsum:
                width = 0.8
                offsets = (np.arange(self.n_scenarios) - (self.n_scenarios - 1) / 2) * width
                ax.bar(self.years + offsets[optim_scenario], 
                       values, 
                       width=width, 
                       label=self.scenario_names[optim_scenario],
                       color = colors[optim_scenario])
                ax1.plot(self.years, 
                        cumsum,
                        label=f'{self.scenario_names[optim_scenario]}',
                        linewidth=4,
                        color = colors[optim_scenario])
                ax1.set_xlabel('Year')
                ax1.set_ylabel(f'Cumulative {self.labels[parameter]}')
                ax1.spines['top'].set_visible(False)
                ax1.spines['right'].set_visible(False)
                ax1.legend(frameon=False)
                fig1.savefig(f'figures\{self.co2_scenario}_Cumulative {parameter}.png',
                dpi=500,
                bbox_inches='tight')
            else:
                ax.plot(self.years, 
                        values, 
                        label=self.scenario_names[optim_scenario], 
                        linewidth=4,
                        color = colors[optim_scenario])
            ax.set_xlabel('Year')
            ax.set_ylabel(self.labels[parameter])
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.legend(frameon=False)
        fig.savefig(f'figures\{self.co2_scenario}_{parameter}.png', 
                dpi=500, 
                bbox_inches='tight')
        if plot_cumsum:
            fig1.savefig(f'figures\{self.co2_scenario}_cumulative {parameter}.png', 
                    dpi=500, 
                    bbox_inches='tight')
        plt.close()
    
    def _add_labels(self, ax, x, y):
        for i in range(len(x)):
            ax.text(i, 
                    y[i] + 0.1, 
                    f'{y[i]:.2f}', 
                    ha='center', 
                    fontsize=10) 

    def plot_obj(self):
        objs = []
        labels = [name.replace(' ', '\n') for name in self.scenario_names.values()]

        for optim_scenario in self.scenario_names.keys():
            wb = xl.load_workbook(self.file_name[optim_scenario]) 
            sheet_name = wb.sheetnames
            df = {sheet: pd.read_excel(self.file_name[optim_scenario], sheet_name=sheet) for sheet in sheet_name}
            objs.append(df['Objective Value'].iloc[0, 0])

        colors = sns.color_palette("tab10", n_colors=len(self.scenario_names.keys())+1)
        fig, ax = plt.subplots(figsize=(8, 4), dpi=600)
        ax.bar(labels, 
               objs, 
               color=colors[:-1],
               width = 0.6)
        self._add_labels(ax, list(self.scenario_names.values()), objs)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylabel('Total Net Cost (Trillion USD$_{2030}$)')
        fig.savefig(f'figures\{self.co2_scenario}_objective.png', 
                    dpi=500, 
                    bbox_inches='tight')

